from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from io import BytesIO
from datetime import datetime
from typing import List

# ─── RANGLAR ────────────────────────────────────────────────────────────────
PRIMARY_HEX   = "2563EB"
HEADER_BG     = "2563EB"
HEADER_FG     = "FFFFFF"
ROW_ALT       = "F1F5F9"
ROW_NORMAL    = "FFFFFF"
BORDER_COLOR  = "E2E8F0"
TOTAL_BG      = "DBEAFE"
TOTAL_FG      = "1E40AF"

# ─── STYLE HELPERLAR ────────────────────────────────────────────────────────
def _header_font(size=10):
    return Font(name="Arial", bold=True, color=HEADER_FG, size=size)

def _normal_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _header_fill():
    return PatternFill("solid", fgColor=HEADER_BG)

def _alt_fill():
    return PatternFill("solid", fgColor=ROW_ALT)

def _total_fill():
    return PatternFill("solid", fgColor=TOTAL_BG)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _right():
    return Alignment(horizontal="right", vertical="center")

def _thin_border():
    thin = Side(style="thin", color=BORDER_COLOR)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _set_header_row(ws, headers: list, row: int = 1, start_col: int = 1):
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=header)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _center()
        cell.border = _thin_border()

def _set_data_row(ws, data: list, row: int, alt: bool = False, start_col: int = 1):
    for i, value in enumerate(data):
        cell = ws.cell(row=row, column=start_col + i, value=value)
        cell.font = _normal_font()
        cell.fill = _alt_fill() if alt else PatternFill("solid", fgColor=ROW_NORMAL)
        cell.alignment = _center()
        cell.border = _thin_border()

def _auto_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, min_width), max_width)


# ════════════════════════════════════════════════════════════════════════════
# 1. OYLIK MOLIYAVIY HISOBOT EXCEL
# ════════════════════════════════════════════════════════════════════════════
def generate_monthly_report_excel(
    clinic_name: str,
    report_month: str,
    payments: List[dict],
    top_services: List[dict],
    daily_revenue: List[dict]
) -> BytesIO:
    wb = Workbook()

    # ── 1-varaq: To'lovlar ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "To'lovlar"

    # Sarlavha
    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value = f"{clinic_name} — Oylik Hisobot: {report_month}"
    title_cell.font = Font(name="Arial", bold=True, size=14, color=PRIMARY_HEX)
    title_cell.alignment = _center()
    title_cell.fill = PatternFill("solid", fgColor="DBEAFE")
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:H2")
    ws1["A2"].value = f"Yaratildi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws1["A2"].font = _normal_font(size=9, color="64748B")
    ws1["A2"].alignment = _center()
    ws1.row_dimensions[2].height = 18

    # Header
    headers = ["Chek №", "Sana", "Bemor", "Kassir", "To'lov usuli",
               "Chegirma", "Jami summa", "Status"]
    _set_header_row(ws1, headers, row=3)
    ws1.row_dimensions[3].height = 22

    # Ma'lumotlar
    total_sum = 0
    for i, p in enumerate(payments, start=4):
        alt = (i % 2 == 0)
        row_data = [
            p.get("receipt_number", ""),
            str(p.get("created_at", ""))[:10],
            p.get("patient_name", ""),
            p.get("cashier_name", ""),
            p.get("payment_method", ""),
            p.get("discount", 0),
            p.get("total_amount", 0),
            p.get("payment_status", "")
        ]
        _set_data_row(ws1, row_data, row=i, alt=alt)
        total_sum += float(p.get("total_amount", 0))
        ws1.row_dimensions[i].height = 18

    # Jami qator
    total_row = len(payments) + 4
    ws1.merge_cells(f"A{total_row}:F{total_row}")
    total_label = ws1[f"A{total_row}"]
    total_label.value = "JAMI / ИТОГО"
    total_label.font = Font(name="Arial", bold=True, size=11, color=TOTAL_FG)
    total_label.fill = _total_fill()
    total_label.alignment = _right()
    total_label.border = _thin_border()

    total_value = ws1[f"G{total_row}"]
    total_value.value = total_sum
    total_value.font = Font(name="Arial", bold=True, size=11, color=TOTAL_FG)
    total_value.fill = _total_fill()
    total_value.alignment = _right()
    total_value.border = _thin_border()
    total_value.number_format = '#,##0'

    ws1.row_dimensions[total_row].height = 24
    _auto_width(ws1)

    # ── 2-varaq: Kunlik daromad ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Kunlik daromad")

    ws2.merge_cells("A1:C1")
    ws2["A1"].value = f"Kunlik daromad — {report_month}"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color=PRIMARY_HEX)
    ws2["A1"].alignment = _center()
    ws2["A1"].fill = PatternFill("solid", fgColor="DBEAFE")
    ws2.row_dimensions[1].height = 28

    _set_header_row(ws2, ["Sana", "Tranzaksiyalar", "Daromad (so'm)"], row=2)
    ws2.row_dimensions[2].height = 22

    for i, d in enumerate(daily_revenue, start=3):
        _set_data_row(ws2, [
            d.get("date", ""),
            d.get("count", 0),
            d.get("total", 0)
        ], row=i, alt=(i % 2 == 0))
        ws2.cell(row=i, column=3).number_format = '#,##0'
        ws2.row_dimensions[i].height = 18

    # Chart — kunlik daromad grafigi
    if daily_revenue:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Kunlik Daromad"
        chart.y_axis.title = "So'm"
        chart.x_axis.title = "Sana"
        chart.style = 10
        chart.width = 18
        chart.height = 12

        data_ref = Reference(ws2, min_col=3, min_row=2, max_row=len(daily_revenue) + 2)
        cats_ref = Reference(ws2, min_col=1, min_row=3, max_row=len(daily_revenue) + 2)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws2.add_chart(chart, f"E2")

    _auto_width(ws2)

    # ── 3-varaq: Top xizmatlar ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Top xizmatlar")

    ws3.merge_cells("A1:D1")
    ws3["A1"].value = f"Eng ko'p talab qilingan xizmatlar — {report_month}"
    ws3["A1"].font = Font(name="Arial", bold=True, size=13, color=PRIMARY_HEX)
    ws3["A1"].alignment = _center()
    ws3["A1"].fill = PatternFill("solid", fgColor="DBEAFE")
    ws3.row_dimensions[1].height = 28

    _set_header_row(ws3, ["№", "Xizmat nomi", "Soni", "Daromad (so'm)"], row=2)
    ws3.row_dimensions[2].height = 22

    for i, s in enumerate(top_services, start=3):
        _set_data_row(ws3, [
            i - 2,
            s.get("name", ""),
            s.get("count", 0),
            s.get("revenue", 0)
        ], row=i, alt=(i % 2 == 0))
        ws3.cell(row=i, column=4).number_format = '#,##0'
        ws3.row_dimensions[i].height = 18

    _auto_width(ws3)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ════════════════════════════════════════════════════════════════════════════
# 2. BEMORLAR RO'YXATI EXCEL
# ════════════════════════════════════════════════════════════════════════════
def generate_patients_excel(clinic_name: str, patients: List[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bemorlar"

    ws.merge_cells("A1:I1")
    ws["A1"].value = f"{clinic_name} — Bemorlar ro'yxati"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=PRIMARY_HEX)
    ws["A1"].alignment = _center()
    ws["A1"].fill = PatternFill("solid", fgColor="DBEAFE")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    ws["A2"].value = f"Yaratildi: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Jami: {len(patients)} ta bemor"
    ws["A2"].font = _normal_font(size=9, color="64748B")
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 16

    headers = ["Kod", "F.I.O", "Tug'ilgan sana", "Jinsi",
               "Telefon", "Qon guruhi", "Allergiya", "Manzil", "Ro'yxatdan o'tgan"]
    _set_header_row(ws, headers, row=3)
    ws.row_dimensions[3].height = 22

    for i, p in enumerate(patients, start=4):
        _set_data_row(ws, [
            p.get("patient_code", ""),
            p.get("full_name", ""),
            str(p.get("birth_date", "")) if p.get("birth_date") else "",
            p.get("gender", ""),
            p.get("phone", ""),
            p.get("blood_type", ""),
            p.get("allergies", ""),
            p.get("address", ""),
            str(p.get("created_at", ""))[:10]
        ], row=i, alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 18

    _auto_width(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ════════════════════════════════════════════════════════════════════════════
# 3. NAVBATLAR RO'YXATI EXCEL
# ════════════════════════════════════════════════════════════════════════════
def generate_appointments_excel(
    clinic_name: str,
    appointments: List[dict],
    report_date: str = ""
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Navbatlar"

    title = f"{clinic_name} — Navbatlar"
    if report_date:
        title += f" ({report_date})"

    ws.merge_cells("A1:G1")
    ws["A1"].value = title
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=PRIMARY_HEX)
    ws["A1"].alignment = _center()
    ws["A1"].fill = PatternFill("solid", fgColor="DBEAFE")
    ws.row_dimensions[1].height = 28

    headers = ["Navbat №", "Sana", "Vaqt", "Bemor", "Shifokor", "Sabab", "Status"]
    _set_header_row(ws, headers, row=2)
    ws.row_dimensions[2].height = 22

    status_colors = {
        "waiting":     "FEF9C3",
        "in_progress": "DBEAFE",
        "completed":   "DCFCE7",
        "cancelled":   "FEE2E2",
        "no_show":     "F3F4F6"
    }

    for i, a in enumerate(appointments, start=3):
        row_data = [
            a.get("queue_number", ""),
            str(a.get("appointment_date", "")),
            str(a.get("appointment_time", "")) if a.get("appointment_time") else "",
            a.get("patient_name", ""),
            a.get("doctor_name", ""),
            a.get("visit_reason", ""),
            a.get("status", "")
        ]
        status = a.get("status", "")
        row_color = status_colors.get(status, ROW_NORMAL)

        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.font = _normal_font()
            cell.fill = PatternFill("solid", fgColor=row_color)
            cell.alignment = _center()
            cell.border = _thin_border()

        ws.row_dimensions[i].height = 18

    _auto_width(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer